import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import json

def gerar_relatorio_excel(dados, caminho_arquivo):
    """
    Gera um relatório Excel formatado a partir de uma lista de dicionários (dados).
    :param dados: Lista de dicionários, cada um representando um item/grupo.
    :param caminho_arquivo: Caminho para salvar o arquivo Excel.
    """
    # Cria DataFrame
    df = pd.DataFrame(dados)
    # Ordena colunas principais se existirem
    colunas_principais = ['Grupo', 'Nome', 'Quantidade', 'Descrição', 'Categoria', 'Local', 'Estado', 'Data de Adição']
    colunas = [c for c in colunas_principais if c in df.columns] + [c for c in df.columns if c not in colunas_principais]
    df = df[colunas]
    # Salva em Excel
    df.to_excel(caminho_arquivo, index=False)
    # Ajusta formatação com openpyxl
    wb = openpyxl.load_workbook(caminho_arquivo)
    ws = wb.active
    # Estilo cabeçalho
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4F81BD')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Ajusta largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 40))
    # Bordas
    thin = Side(border_style="thin", color="CCCCCC")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    # Congela cabeçalho
    ws.freeze_panes = 'A2'
    # Salva
    wb.save(caminho_arquivo)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python relatorio_excel.py <entrada.json> <saida.xlsx>")
        sys.exit(1)
    arquivo_json = sys.argv[1]
    arquivo_excel = sys.argv[2]
    with open(arquivo_json, 'r', encoding='utf-8') as f:
        dados = json.load(f)
    gerar_relatorio_excel(dados, arquivo_excel)
    print("Relatório gerado com sucesso!")
